"""
data/plantilla.py — Genera plantilla Excel con DOS hojas de datos:
  • "Base"    — período de referencia para ajustar el modelo (LBEn)
  • "Reporte" — período de evaluación del desempeño energético (opcional)

Soporta tres frecuencias:
  • mensual: fechas en formato "ene-2024"
  • diario:  fechas en formato "dd/mm/yyyy"
  • horario: fechas en formato "dd/mm/yyyy HH:00"

Alineado con Resolución UPME 16 de 2024 — Metodología LBEn.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
from typing import List, Union

# ── Paleta de colores institucional ──────────────────────────────────────────
# Azul UPME oscuro  / Azul UPME medio / Verde consumo / Violeta variables
_C_UPME_DARK   = "1B3A6B"
_C_UPME_MID    = "2563A8"
_C_UPME_LIGHT  = "EBF2FB"
_C_GREEN       = "1A6B3C"
_C_GREEN_LIGHT = "E9F7EF"
_C_VIOLET      = "4A235A"
_C_VIOLET_LIGHT= "F5EEF8"
_C_AMBER       = "7D5A0A"
_C_AMBER_LIGHT = "FEF9E7"
_C_ROW_ALT     = "F4F6FA"
_C_ROW_NORM    = "FFFFFF"
_C_REPORT_HDR  = "6B3A00"
_C_REPORT_CELL = "FFF8EE"
_C_DIAS_LIGHT  = "DDE8F5"   # azul claro para celdas de días

# ── Estilos base ──────────────────────────────────────────────────────────────
_THIN  = Side(style="thin",   color="C5D0E0")
_MED   = Side(style="medium", color="2563A8")
_B     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_B_MED = Border(left=_MED,  right=_MED,  top=_MED,  bottom=_MED)
_AC    = Alignment(horizontal="center", vertical="center", wrap_text=False)
_AL    = Alignment(horizontal="left",   vertical="center", wrap_text=False)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)


# ── Formateo de fechas ────────────────────────────────────────────────────────
_MESES_ES = ["ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]

def _fmt_fecha(f: Union[date, datetime], frecuencia: str) -> str:
    if frecuencia == "mensual":
        return f"{_MESES_ES[f.month-1]}-{f.year}"
    if frecuencia == "diario":
        return f.strftime("%d/%m/%Y")
    if frecuencia == "horario":
        return f.strftime("%d/%m/%Y %H:00")
    return str(f)

def _clave_fecha(valor, frecuencia: str) -> str:
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return _fmt_fecha(valor, frecuencia).lower()
    if isinstance(valor, date) and not isinstance(valor, datetime):
        return _fmt_fecha(valor, frecuencia).lower()
    return str(valor).strip().lower()

def _freq_text(frecuencia: str) -> str:
    return {"mensual": "Mensual (mes-año)", "diario": "Diario (dd/mm/yyyy)",
            "horario": "Horario (dd/mm/yyyy HH:00)"}.get(frecuencia, frecuencia)

def _ancho_col(j: int, frecuencia: str, col_name: str = "") -> int:
    if j == 1:
        return {"mensual": 14, "diario": 14, "horario": 20}.get(frecuencia, 16)
    if "Días_facturación" in col_name or "dias" in col_name.lower():
        return 18
    if "Ajuste" in col_name or "NR" in col_name:
        return 32
    return 18


# ── Escritura de encabezado de banda ─────────────────────────────────────────
def _titulo_banda(ws, texto: str, n_cols: int, row: int,
                  bg: str = _C_UPME_DARK, fg: str = "FFFFFF", size: int = 11):
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws.cell(row=row, column=1, value=texto)
    c.fill      = _fill(bg)
    c.font      = _font(bold=True, color=fg, size=size)
    c.alignment = _AC
    c.border    = _B_MED
    ws.row_dimensions[row].height = 22


# ── API pública ───────────────────────────────────────────────────────────────

def generar_plantilla(path, modelo_id, col_consumo, vars_independientes,
                      fechas_base=None, fechas_reporte=None,
                      nombre_proyecto="", zona_climatica="", unidad="kWh",
                      frecuencia="mensual"):
    wb = openpyxl.Workbook()
    _hoja_instrucciones(wb, modelo_id, nombre_proyecto, zona_climatica, unidad,
                        col_consumo, vars_independientes,
                        fechas_base, fechas_reporte, frecuencia)
    _hoja_datos(wb, "Base", col_consumo, vars_independientes,
                fechas_base or [], unidad,
                hdr_fecha_bg=_C_UPME_DARK, hdr_fecha_fg="FFFFFF",
                cell_fecha_bg=_C_UPME_LIGHT, cell_fecha_fg=_C_UPME_DARK,
                titulo="Período de línea base",
                incluir_anr=True,
                incluir_dias=True,
                frecuencia=frecuencia)
    if fechas_reporte:
        _hoja_datos(wb, "Reporte", col_consumo, vars_independientes,
                    fechas_reporte, unidad,
                    hdr_fecha_bg=_C_REPORT_HDR, hdr_fecha_fg="FFFFFF",
                    cell_fecha_bg=_C_REPORT_CELL, cell_fecha_fg=_C_REPORT_HDR,
                    titulo="Período de evaluación del desempeño energético",
                    incluir_anr=False,
                    incluir_dias=True,
                    frecuencia=frecuencia)
    wb.active = wb["Base"]
    wb.save(path)


def expandir_reporte(path_origen: str, path_destino: str,
                     nuevas_fechas: list,
                     col_consumo: str,
                     vars_independientes: list,
                     unidad: str = "kWh",
                     frecuencia: str = "mensual"):
    """
    Lee el Excel base y agrega nuevas_fechas a la hoja 'Reporte',
    sin duplicar períodos ya existentes.
    Retorna (n_existentes, n_agregadas).
    """
    import shutil, os, tempfile

    wb = openpyxl.load_workbook(path_origen)

    fechas_existentes_clave: set = set()
    filas_existentes: list = []

    if "Reporte" in wb.sheetnames:
        ws_rep = wb["Reporte"]
        # +1 por columna Días_facturación
        n_cols = 1 + 1 + 1 + len(vars_independientes)
        for row in ws_rep.iter_rows(min_row=3, max_col=n_cols, values_only=True):
            fecha_val = row[0]
            if fecha_val is None or str(fecha_val).strip() == "":
                continue
            clave = _clave_fecha(fecha_val, frecuencia)
            if not clave:
                continue
            fecha_str = (_fmt_fecha(fecha_val, frecuencia)
                         if isinstance(fecha_val, (datetime, date))
                         else str(fecha_val).strip())
            fechas_existentes_clave.add(clave)
            filas_existentes.append((fecha_str, list(row[1:])))

    fechas_a_agregar = [
        f for f in nuevas_fechas
        if _fmt_fecha(f, frecuencia).lower() not in fechas_existentes_clave
    ]
    n_existentes = len(filas_existentes)
    n_agregadas  = len(fechas_a_agregar)

    if n_agregadas == 0:
        wb.close()
        if os.path.normpath(path_destino) != os.path.normpath(path_origen):
            shutil.copy2(path_origen, path_destino)
        return n_existentes, 0

    # Reconstruir hoja Reporte
    if "Reporte" in wb.sheetnames:
        del wb["Reporte"]

    # Período | Consumo | Días_facturación | Variables
    columnas   = ["Período", col_consumo, "Días_facturación"] + vars_independientes
    n_cols_tot = len(columnas)

    ws = wb.create_sheet("Reporte")

    # Banda de título
    _titulo_banda(ws,
                  "Período de evaluación del desempeño energético",
                  n_cols_tot, 1,
                  bg=_C_REPORT_HDR)

    # Encabezados columnas (fila 2)
    hdr_fills = (
        [_fill(_C_REPORT_HDR), _fill(_C_GREEN), _fill(_C_UPME_MID)]
        + [_fill(_C_VIOLET)] * len(vars_independientes)
    )
    for j, (col, fill) in enumerate(zip(columnas, hdr_fills), start=1):
        c = ws.cell(row=2, column=j, value=col)
        c.fill      = fill
        c.font      = _font(bold=True, color="FFFFFF", size=10)
        c.alignment = _AC
        c.border    = _B
        ws.column_dimensions[get_column_letter(j)].width = _ancho_col(j, frecuencia, col)

    # Filas existentes
    for i, (fecha_str, valores) in enumerate(filas_existentes):
        r  = i + 3
        bg = _C_ROW_ALT if i % 2 == 0 else _C_ROW_NORM
        cf = ws.cell(row=r, column=1, value=fecha_str)
        cf.fill = _fill(_C_REPORT_CELL); cf.font = _font(bold=True, color=_C_REPORT_HDR)
        cf.alignment = _AC; cf.border = _B
        for j_off, val in enumerate(valores):
            c = ws.cell(row=r, column=j_off + 2, value=val)
            col_name = columnas[j_off + 1]
            if "Días_facturación" in col_name:
                c.fill = _fill(_C_DIAS_LIGHT)
                c.font = _font(color=_C_UPME_DARK)
                c.number_format = "0"
            else:
                c.fill = _fill(bg)
                if val is not None:
                    c.number_format = "#,##0.00"
            c.alignment = _AC
            c.border = _B

    # Filas nuevas (vacías)
    base_i = len(filas_existentes)
    for k, fecha_obj in enumerate(fechas_a_agregar):
        i  = base_i + k
        r  = i + 3
        bg = _C_ROW_ALT if i % 2 == 0 else _C_ROW_NORM
        cf = ws.cell(row=r, column=1, value=_fmt_fecha(fecha_obj, frecuencia))
        cf.fill = _fill(_C_REPORT_CELL); cf.font = _font(bold=True, color=_C_REPORT_HDR)
        cf.alignment = _AC; cf.border = _B
        for j in range(2, n_cols_tot + 1):
            c = ws.cell(row=r, column=j)
            col_name = columnas[j - 1]
            if "Días_facturación" in col_name:
                c.fill          = _fill(_C_DIAS_LIGHT)
                c.font          = _font(color=_C_UPME_DARK)
                c.number_format = "0"
                c.value         = 30
            else:
                c.fill          = _fill(bg)
                c.number_format = "#,##0.00"
            c.alignment = _AC
            c.border    = _B

    ws.freeze_panes = "B3"

    # Guardar de forma atómica
    fd, tmp_path = tempfile.mkstemp(
        suffix=".xlsx", dir=os.path.dirname(path_destino) or ".")
    os.close(fd)
    try:
        wb.save(tmp_path)
        wb.close()
        if os.path.exists(path_destino):
            os.remove(path_destino)
        shutil.move(tmp_path, path_destino)
    except Exception:
        if os.path.exists(tmp_path):
            try: os.remove(tmp_path)
            except Exception: pass
        raise

    return n_existentes, n_agregadas


# ── Helpers internos ──────────────────────────────────────────────────────────

def _hoja_instrucciones(wb, modelo_id, nombre_proyecto, zona_climatica, unidad,
                        col_consumo, vars_ind, fh, fr, frecuencia):
    ws = wb.active
    ws.title = "Instrucciones"
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50
    ws.sheet_view.showGridLines = False

    # ── Cabecera institucional ─────────────────────────────────────────────
    ws.merge_cells("A1:B1")
    ws["A1"] = "SISTEMA DE GESTIÓN LBEn — UPME"
    ws["A1"].fill      = _fill(_C_UPME_DARK)
    ws["A1"].font      = _font(bold=True, color="FFFFFF", size=13)
    ws["A1"].alignment = _AC
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:B2")
    ws["A2"] = "Línea Base de Consumo Energético (LBEn) — Resolución UPME 16 de 2024"
    ws["A2"].fill      = _fill(_C_UPME_MID)
    ws["A2"].font      = _font(bold=True, color="FFFFFF", size=11)
    ws["A2"].alignment = _AC
    ws.row_dimensions[2].height = 22

    # ── Metadatos del proyecto ─────────────────────────────────────────────
    meta = [
        ("Proyecto:",        nombre_proyecto or "—"),
        ("Zona climática:",  zona_climatica  or "—"),
        ("Unidad energética:", unidad),
        ("Frecuencia de datos:", _freq_text(frecuencia)),
        ("Modelo LBEn:",     _modelo_label(modelo_id)),
        ("Variable de consumo:", col_consumo),
        ("Variables relevantes:", ", ".join(vars_ind) if vars_ind else "Ninguna"),
        ("Período base:",
         f"{_fmt_fecha(fh[0], frecuencia)} — {_fmt_fecha(fh[-1], frecuencia)}  ({len(fh)} períodos)"
         if fh else "—"),
        ("Período de evaluación:",
         f"{_fmt_fecha(fr[0], frecuencia)} — {_fmt_fecha(fr[-1], frecuencia)}  ({len(fr)} períodos)"
         if fr else "No configurado"),
    ]
    for i, (k, v) in enumerate(meta, start=4):
        ws.row_dimensions[i].height = 18
        ka = ws.cell(row=i, column=1, value=k)
        ka.font = _font(bold=True, color=_C_UPME_DARK, size=10)
        ka.fill = _fill(_C_UPME_LIGHT)
        ka.border = _B
        ka.alignment = _AL
        va = ws.cell(row=i, column=2, value=v)
        va.font = _font(size=10)
        va.border = _B
        va.alignment = _AL

    # ── Instrucciones ──────────────────────────────────────────────────────
    row_ini = 4 + len(meta) + 1
    ws.merge_cells(f"A{row_ini}:B{row_ini}")
    ws.cell(row=row_ini, column=1,
            value="INSTRUCCIONES PARA EL DILIGENCIAMIENTO").font = _font(bold=True, color="FFFFFF", size=10)
    ws.cell(row=row_ini, column=1).fill = _fill(_C_UPME_MID)
    ws.cell(row=row_ini, column=1).alignment = _AC
    ws.row_dimensions[row_ini].height = 20

    instrucciones = [
        ("1.", "Complete la hoja «Base» con los consumos energéticos del período de línea base."),
        ("2.", f"La columna «Período» usa formato {_fmt_periodo_hint(frecuencia)}."),
        ("3.", f"Ingrese los valores numéricos de consumo en {unidad} en la columna de consumo."),
        ("3b.", "Columna «Días_facturación»: ingrese los días reales del período de facturación. "
                "El sistema normalizará automáticamente a 30 días (Ec. 1, Resolución UPME 16/2024). "
                "Por defecto se precargan 30 días — modifique solo si el ciclo fue diferente."),
        ("4.", "Columna «Ajuste_NR»: escriba el motivo si un período es anómalo "
               "(ej.: «mantenimiento», «huelga», «remodelación»). Deje en blanco si el período es normal."),
        ("5.", "Si configuró el período de evaluación, complete la hoja «Reporte» "
               "con los datos del período posterior a las medidas de eficiencia."),
        ("6.", "No elimine ni renombre columnas — el sistema las utiliza por nombre exacto."),
        ("7.", "No deje celdas numéricas vacías dentro del rango de datos."),
        ("8.", "Consulte la Resolución UPME 16 de 2024 para la metodología completa de cálculo de LBEn."),
    ]
    for j, (num, txt) in enumerate(instrucciones, start=row_ini + 1):
        ws.row_dimensions[j].height = 28
        cn = ws.cell(row=j, column=1, value=num)
        cn.font = _font(bold=True, color=_C_UPME_DARK, size=10)
        cn.fill = _fill(_C_UPME_LIGHT if j % 2 == 0 else _C_ROW_NORM)
        cn.border = _B; cn.alignment = _AC
        ct = ws.cell(row=j, column=2, value=txt)
        ct.font = _font(size=10)
        ct.fill = _fill(_C_UPME_LIGHT if j % 2 == 0 else _C_ROW_NORM)
        ct.border = _B
        ct.alignment = Alignment(horizontal="left", vertical="center",
                                  wrap_text=True)


def _modelo_label(modelo_id: str) -> str:
    return {"promedio":  "Valor absoluto de energía (Art. 7.4.1)",
            "cociente":  "Cociente de valores medidos (Art. 7.4.2)",
            "regresion": "Modelo estadístico — Regresión lineal (Art. 7.4.3)"
            }.get(modelo_id, modelo_id.title())


def _fmt_periodo_hint(frecuencia: str) -> str:
    return {"mensual": "«mes-año» (ej.: ene-2024)",
            "diario":  "«dd/mm/yyyy» (ej.: 01/04/2024)",
            "horario": "«dd/mm/yyyy HH:00» (ej.: 01/04/2024 08:00)"
            }.get(frecuencia, "fecha")


def _hoja_datos(wb, nombre_hoja, col_consumo, vars_ind, fechas,
                unidad,
                hdr_fecha_bg, hdr_fecha_fg,
                cell_fecha_bg, cell_fecha_fg,
                titulo,
                incluir_anr=False,
                incluir_dias=False,
                frecuencia="mensual"):
    ws = wb.create_sheet(nombre_hoja)
    ws.sheet_view.showGridLines = False

    # Columnas según Resolución UPME 16/2024:
    # Período | Consumo energético normalizado | [Días_facturación] | Variables | [Ajuste_NR]
    col_consumo_label = f"{col_consumo}"
    columnas = ["Período", col_consumo_label]
    if incluir_dias:
        columnas += ["Días_facturación"]
    columnas += vars_ind
    if incluir_anr:
        columnas += ["Ajuste_NR"]

    n_cols = len(columnas)

    # Colores por tipo de columna
    hdr_fills = (
        [_fill(hdr_fecha_bg), _fill(_C_GREEN)]
        + ([_fill(_C_UPME_MID)] if incluir_dias else [])
        + [_fill(_C_VIOLET)] * len(vars_ind)
        + ([_fill(_C_AMBER)] if incluir_anr else [])
    )
    hdr_fgs = (
        [hdr_fecha_fg, "FFFFFF"]
        + (["FFFFFF"] if incluir_dias else [])
        + ["FFFFFF"] * len(vars_ind)
        + (["FFFFFF"] if incluir_anr else [])
    )

    # Fila 1 — Banda de título
    _titulo_banda(ws, titulo, n_cols, 1, bg=_C_UPME_DARK)

    # Fila 2 — Encabezados de columna
    for j, (col, fill, fg) in enumerate(zip(columnas, hdr_fills, hdr_fgs), start=1):
        c = ws.cell(row=2, column=j, value=col)
        c.fill      = fill
        c.font      = _font(bold=True, color=fg, size=10)
        c.alignment = _AC
        c.border    = _B
        ws.column_dimensions[get_column_letter(j)].width = _ancho_col(j, frecuencia, col)
    ws.row_dimensions[2].height = 22

    # Filas de datos (desde fila 3)
    for i, f in enumerate(fechas):
        r  = i + 3
        bg = _C_ROW_ALT if i % 2 == 0 else _C_ROW_NORM
        ws.row_dimensions[r].height = 18

        for j in range(1, n_cols + 1):
            c = ws.cell(row=r, column=j)
            c.border    = _B
            c.alignment = _AC
            col_name    = columnas[j - 1]

            if j == 1:
                # Columna Período
                c.value     = _fmt_fecha(f, frecuencia)
                c.fill      = _fill(cell_fecha_bg)
                c.font      = _font(bold=True, color=cell_fecha_fg, size=10)

            elif "Días_facturación" in col_name:
                # Columna días de facturación — entero, precargado con 30
                c.fill          = _fill(_C_DIAS_LIGHT)
                c.font          = _font(color=_C_UPME_DARK, size=10)
                c.number_format = "0"
                c.value         = 30

            elif "Ajuste_NR" in col_name or "ajuste" in col_name.lower():
                # Columna ANR — texto libre
                c.fill      = _fill(_C_AMBER_LIGHT)
                c.font      = _font(color=_C_AMBER, size=10)
                c.alignment = _AL
                c.value     = ""

            else:
                # Columnas numéricas (consumo y variables)
                c.fill          = _fill(bg)
                c.font          = _font(color="000000", size=10)
                c.number_format = "#,##0.00"

    ws.freeze_panes = "B3"

    # Borde exterior de la tabla
    _aplicar_borde_exterior(ws, 1, 1, len(fechas) + 2, n_cols)


def _aplicar_borde_exterior(ws, row_ini, col_ini, row_fin, col_fin):
    """Aplica borde medium en el contorno exterior de un rango."""
    borde_ext = Side(style="medium", color=_C_UPME_MID)
    for r in range(row_ini, row_fin + 1):
        for c in range(col_ini, col_fin + 1):
            cell = ws.cell(row=r, column=c)
            left   = borde_ext if c == col_ini  else _THIN
            right  = borde_ext if c == col_fin   else _THIN
            top    = borde_ext if r == row_ini   else _THIN
            bottom = borde_ext if r == row_fin   else _THIN
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)