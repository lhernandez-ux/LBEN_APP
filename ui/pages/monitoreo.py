"""
ui/pages/monitoreo.py
=====================
Flujo de monitoreo:

  1. El usuario selecciona un proyecto guardado (izquierda)
  2. Panel derecho muestra info del proyecto
  3. Sección "Actualizar reporte":
       a. Preview de cuántos períodos nuevos quiere agregar, ADAPTADO a la
          frecuencia del proyecto:
            • mensual → campo "N meses"
            • diario  → campos "desde" / "hasta" (dd/mm/yyyy)
            • horario → campos "desde" / "hasta" + rango horario (HH:00)
          NOTA: el modo diario/horario solo aplica al modelo estadístico
                (regresión). Los modelos promedio/cociente son siempre mensuales.
       b. Botón "Descargar Excel actualizado" → genera filas vacías al final
          de la hoja 'Reporte' y guarda el archivo donde el usuario elija.
       c. El usuario llena los nuevos valores y lo sube.
       d. Al subir: reemplaza el Excel del proyecto y recalcula todo.
"""

import os
import re
from datetime import date, datetime, timedelta
import customtkinter as ctk
from tkinter import filedialog
from ui.theme import COLORS, FONTS, SIZES

_MESES_ABREV = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
}
_MESES_ABREV_INV = {v: k for k, v in _MESES_ABREV.items()}


# ── Helpers de frecuencia ──────────────────────────────────────────────────────

def _frecuencia_proyecto(p: dict) -> str:
    modelo_id = p.get("modelo_id", "promedio")
    if modelo_id != "regresion":
        return "mensual"

    # 1. Usar la frecuencia guardada en el JSON (proyectos nuevos)
    freq = p.get("frecuencia", "").strip().lower()
    if freq in ("mensual", "diario", "horario"):
        return freq

    # 2. Fallback: inferir desde el Excel
    ruta = p.get("ruta_excel", "")
    if ruta and os.path.exists(ruta):
        try:
            import openpyxl
            import re as _re
            from datetime import datetime, date
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            hoja = "Base" if "Base" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[hoja]
            muestra = None
            for row in ws.iter_rows(min_row=4, max_col=1, max_row=10, values_only=True):
                if row[0] is not None:
                    muestra = row[0]
                    break
            wb.close()

            if muestra is not None:
                # Caso 1: openpyxl devolvió datetime nativo
                if isinstance(muestra, datetime):
                    # datetime con hora → horario; solo fecha → diario
                    if muestra.hour != 0 or muestra.minute != 0:
                        return "horario"
                    else:
                        return "diario"
                # Caso 2: openpyxl devolvió date nativo (sin hora)
                if isinstance(muestra, date):
                    return "diario"
                # Caso 3: string — detectar por formato
                s = str(muestra).strip()
                if _re.match(r"^\d{2}/\d{2}/\d{4} \d{2}:00$", s):
                    return "horario"
                if _re.match(r"^\d{2}/\d{2}/\d{4}$", s):
                    return "diario"
                # Si es "ene-2025" u otro formato mensual → mensual
        except Exception:
            pass

    return "mensual"


def _parsear_fecha_excel(valor, frecuencia: str):
    """
    Convierte un valor leído de la hoja Reporte (puede ser str o datetime/date)
    al tipo nativo correspondiente a la frecuencia.
    Retorna date para mensual/diario, datetime para horario. None si falla.
    """
    if valor is None:
        return None
    s = str(valor).strip()
    if not s:
        return None

    if frecuencia == "mensual":
        m = re.match(r"^([a-záéíóúñ]{3})-(\d{4})$", s.lower())
        if m:
            mes_n = _MESES_ABREV.get(m.group(1), 0)
            if mes_n:
                return date(int(m.group(2)), mes_n, 1)
        return None

    if frecuencia == "diario":
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        if hasattr(valor, "date"):
            return valor.date()
        return None

    if frecuencia == "horario":
        for fmt in ("%d/%m/%Y %H:00", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        if isinstance(valor, datetime):
            return valor
        return None

    return None


def _fmt_fecha(f, frecuencia: str) -> str:
    """Formatea para mostrar en la UI y como valor en Excel."""
    if frecuencia == "mensual":
        ab = ["ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]
        return f"{ab[f.month-1]}-{f.year}"
    if frecuencia == "diario":
        return f.strftime("%d/%m/%Y")
    if frecuencia == "horario":
        return f.strftime("%d/%m/%Y %H:00")
    return str(f)


def _ultimo_periodo_reporte(ruta: str, frecuencia: str):
    """
    Lee la hoja Reporte y devuelve el último período existente parseado.
    Retorna date (mensual/diario) o datetime (horario), o None.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        if "Reporte" not in wb.sheetnames:
            wb.close()
            return None
        ws  = wb["Reporte"]
        ult = None
        for row in ws.iter_rows(min_row=4, max_col=1, values_only=True):
            if row[0] is not None and str(row[0]).strip():
                ult = row[0]
        wb.close()
        return _parsear_fecha_excel(ult, frecuencia)
    except Exception:
        return None


def _ultimo_periodo_base(ruta: str, frecuencia: str):
    """
    Lee la hoja Base y devuelve el último período.
    """
    try:
        from data.lector_excel import leer_excel
        df = leer_excel(ruta, hoja="Base")
        if len(df) == 0:
            return None
        val = df.iloc[-1, 0]
        if isinstance(val, datetime):
            if frecuencia == "horario":
                return val
            return val.date()
        if isinstance(val, date):
            return val
        return _parsear_fecha_excel(val, frecuencia)
    except Exception:
        return None


def _generar_fechas_desde(ultimo, n: int, frecuencia: str):
    """
    Genera n períodos consecutivos a partir del siguiente al período `ultimo`.
    """
    fechas = []
    if frecuencia == "mensual":
        año, mes = ultimo.year, ultimo.month
        for _ in range(n):
            mes += 1
            if mes > 12:
                mes = 1; año += 1
            fechas.append(date(año, mes, 1))

    elif frecuencia == "diario":
        cursor = ultimo + timedelta(days=1)
        for _ in range(n):
            fechas.append(cursor)
            cursor += timedelta(days=1)

    elif frecuencia == "horario":
        cursor = ultimo + timedelta(hours=1)
        for _ in range(n):
            fechas.append(cursor)
            cursor += timedelta(hours=1)

    return fechas


def _generar_fechas_rango(desde, hasta, frecuencia: str):
    """
    Genera todos los períodos entre desde y hasta (inclusive).
    `desde` y `hasta` son date (diario) o datetime (horario).
    """
    fechas = []
    cursor = desde
    if frecuencia == "diario":
        while cursor <= hasta:
            fechas.append(cursor)
            cursor += timedelta(days=1)
    elif frecuencia == "horario":
        while cursor <= hasta:
            fechas.append(cursor)
            cursor += timedelta(hours=1)
    return fechas


def _parsear_date_str(s: str) -> date:
    """Parsea dd/mm/yyyy → date. Lanza ValueError si no puede."""
    return datetime.strptime(s.strip(), "%d/%m/%Y").date()


def _parsear_datetime_str(fecha_s: str, hora_s: str) -> datetime:
    """Parsea 'dd/mm/yyyy' + 'HH' → datetime. Lanza ValueError si no puede."""
    dt = datetime.strptime(fecha_s.strip(), "%d/%m/%Y")
    hora = int(hora_s.strip())
    return dt.replace(hour=hora)


# ─────────────────────────────────────────────────────────────────────────────

class MonitoreoPage(ctk.CTkFrame):
    def __init__(self, parent, app):
        super().__init__(parent, fg_color=COLORS.bg_main)
        self.app = app
        self._proyecto_sel = None
        self._resultado    = None
        self._build()

    def on_enter(self, **kwargs):
        self._refrescar_lista()

    # ═════════════════════════════════════════════════════════════════════════
    # ESTRUCTURA
    # ═════════════════════════════════════════════════════════════════════════
    def _build(self):
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        hdr = ctk.CTkFrame(self, fg_color=COLORS.bg_card, corner_radius=0, height=60)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.grid_propagate(False)
        hdr.grid_columnconfigure(1, weight=1)

        ctk.CTkButton(hdr, text="← Inicio", width=80, height=32,
                      fg_color="transparent", hover_color=COLORS.bg_main,
                      text_color=COLORS.text_secondary,
                      font=(FONTS.family, FONTS.size_sm),
                      command=lambda: self.app.show_page("IntroPage")
                      ).grid(row=0, column=0, padx=16)

        ctk.CTkLabel(hdr, text="📡  Monitoreo de proyectos",
                     font=(FONTS.family, FONTS.size_lg, "bold"),
                     text_color=COLORS.text_primary
                     ).grid(row=0, column=1, sticky="w", padx=4)

        body = ctk.CTkFrame(self, fg_color="transparent")
        body.grid(row=1, column=0, sticky="nsew", padx=16, pady=12)
        body.grid_rowconfigure(0, weight=1)
        body.grid_columnconfigure(0, minsize=270, weight=0)
        body.grid_columnconfigure(1, weight=1)

        lista_card = ctk.CTkFrame(body, fg_color=COLORS.bg_card,
                                   corner_radius=10, border_width=1,
                                   border_color=COLORS.border)
        lista_card.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        lista_card.grid_rowconfigure(1, weight=1)
        lista_card.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(lista_card, text="Proyectos guardados",
                     font=(FONTS.family, FONTS.size_sm, "bold"),
                     text_color=COLORS.primary
                     ).grid(row=0, column=0, sticky="w", padx=14, pady=(14, 6))

        self.lista_sv = ctk.CTkScrollableFrame(lista_card, fg_color="transparent")
        self.lista_sv.grid(row=1, column=0, sticky="nsew", padx=6, pady=(0, 8))
        self.lista_sv.grid_columnconfigure(0, weight=1)

        self.right_sv = ctk.CTkScrollableFrame(body, fg_color="transparent")
        self.right_sv.grid(row=0, column=1, sticky="nsew")
        self.right_sv.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(self.right_sv,
                     text="📋\n\nSelecciona un proyecto de la lista",
                     font=(FONTS.family, FONTS.size_md),
                     text_color=COLORS.text_secondary,
                     justify="center").pack(expand=True, pady=100)

    # ═════════════════════════════════════════════════════════════════════════
    # LISTA
    # ═════════════════════════════════════════════════════════════════════════
    def _refrescar_lista(self):
        for w in self.lista_sv.winfo_children():
            w.destroy()
        self._limpiar_derecha()

        from data.gestor_proyectos import listar_proyectos
        proyectos = listar_proyectos()
        if not proyectos:
            ctk.CTkLabel(self.lista_sv,
                         text="Sin proyectos.\nCrea uno desde\n'Comenzar →'",
                         font=(FONTS.family, FONTS.size_sm),
                         text_color=COLORS.text_secondary,
                         justify="center").pack(pady=40)
            return
        for p in proyectos:
            self._card_proyecto(p)

    def _card_proyecto(self, p: dict):
        card = ctk.CTkFrame(self.lista_sv, fg_color=COLORS.bg_main,
                             corner_radius=8, border_width=1,
                             border_color=COLORS.border, cursor="hand2")
        card.pack(fill="x", padx=4, pady=3)

        fila_top = ctk.CTkFrame(card, fg_color="transparent")
        fila_top.pack(fill="x", padx=6, pady=(6, 0))
        fila_top.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(fila_top, text=p.get("nombre_proyecto", "Sin nombre"),
                     font=(FONTS.family, FONTS.size_xs, "bold"),
                     text_color=COLORS.primary, anchor="w"
                     ).grid(row=0, column=0, sticky="w", padx=4)

        btn_del = ctk.CTkButton(fila_top, text="✕", width=22, height=22,
                                fg_color="transparent",
                                hover_color="#FADBD8",
                                text_color=COLORS.text_secondary,
                                font=(FONTS.family, FONTS.size_xs, "bold"),
                                corner_radius=4,
                                command=lambda proj=p: self._confirmar_eliminar(proj))
        btn_del.grid(row=0, column=1, padx=(2, 2))

        freq = _frecuencia_proyecto(p)
        freq_label = {"mensual": "Mensual", "diario": "Diario", "horario": "Horario"}.get(freq, freq)
        ctk.CTkLabel(card,
                     text=f"{p.get('modelo_id','').title()}  ·  {freq_label}  ·  {p.get('guardado_en','')[:10]}",
                     font=(FONTS.family, FONTS.size_xs),
                     text_color=COLORS.text_secondary, anchor="w"
                     ).pack(fill="x", padx=10, pady=(0, 8))

        def _click(e, proj=p):
            self._seleccionar(proj)
        def _on(e):  card.configure(fg_color=COLORS.primary_light)
        def _off(e): card.configure(fg_color=COLORS.bg_main)
        bind_widgets = [card] + [w for w in card.winfo_children()
                                  if w != btn_del] + list(fila_top.winfo_children())[:-1]
        for w in bind_widgets:
            w.bind("<Button-1>", _click)
            w.bind("<Enter>", _on); w.bind("<Leave>", _off)

    def _confirmar_eliminar(self, p: dict):
        from tkinter import messagebox
        nombre = p.get("nombre_proyecto", "este proyecto")
        confirmar = messagebox.askyesno(
            title="Eliminar proyecto",
            message=f"¿Seguro que quieres eliminar '{nombre}'?\n\nEsto solo borra el registro guardado.\nEl archivo Excel NO se elimina.",
            icon="warning",
        )
        if confirmar:
            from data.gestor_proyectos import eliminar_proyecto
            eliminar_proyecto(p["_archivo"])
            self._refrescar_lista()

    # ═════════════════════════════════════════════════════════════════════════
    # PANEL DERECHO
    # ═════════════════════════════════════════════════════════════════════════
    def _limpiar_derecha(self):
        for w in self.right_sv.winfo_children():
            w.destroy()
        self._resultado = None

    def _seleccionar(self, p: dict):
        self._proyecto_sel = p
        self._limpiar_derecha()
        self._panel_info(p)
        self._panel_actualizar(p)
        ctk.CTkFrame(self.right_sv, fg_color=COLORS.border, height=2
                     ).pack(fill="x", padx=4, pady=(8, 0))

    # ── Info del proyecto ─────────────────────────────────────────────────────
    def _panel_info(self, p: dict):
        sv = self.right_sv
        ctk.CTkLabel(sv, text=p.get("nombre_proyecto", "Proyecto"),
                     font=(FONTS.family, FONTS.size_xl, "bold"),
                     text_color=COLORS.primary, anchor="w"
                     ).pack(fill="x", padx=4, pady=(4, 6))

        info = ctk.CTkFrame(sv, fg_color=COLORS.primary_light,
                             corner_radius=8, border_width=1,
                             border_color=COLORS.border)
        info.pack(fill="x", padx=4, pady=(0, 8))
        info.grid_columnconfigure((0, 1, 2), weight=1)

        freq = _frecuencia_proyecto(p)
        freq_label = {"mensual": "Mensual", "diario": "Diario", "horario": "Horario"}.get(freq, freq.title())
        campos = [
            ("Modelo",            p.get("modelo_id", "—").title()),
            ("Frecuencia",        freq_label),
            ("Unidad",            p.get("unidad_energia", "—")),
            ("Período base", p.get("periodo_base", "—")),
            ("Columna consumo",   p.get("col_consumo", "—")),
            ("Guardado",          p.get("guardado_en", "")[:10]),
        ]
        for i, (lbl, val) in enumerate(campos):
            r, c = divmod(i, 3)
            f = ctk.CTkFrame(info, fg_color="transparent")
            f.grid(row=r, column=c, sticky="ew", padx=12, pady=6)
            ctk.CTkLabel(f, text=lbl,
                         font=(FONTS.family, FONTS.size_xs),
                         text_color=COLORS.text_secondary, anchor="w").pack(anchor="w")
            ctk.CTkLabel(f, text=val,
                         font=(FONTS.family, FONTS.size_sm, "bold"),
                         text_color=COLORS.primary, anchor="w").pack(anchor="w")

    # ── Panel de actualización — despacha según frecuencia ────────────────────
    def _panel_actualizar(self, p: dict):
        freq = _frecuencia_proyecto(p)
        if freq == "mensual":
            self._panel_actualizar_mensual(p)
        elif freq == "diario":
            self._panel_actualizar_diario(p)
        elif freq == "horario":
            self._panel_actualizar_horario(p)

    # ══════════════════════════════════════════════════════════════════════════
    # PANEL MENSUAL
    # ══════════════════════════════════════════════════════════════════════════
    def _panel_actualizar_mensual(self, p: dict):
        sv   = self.right_sv
        ruta = p.get("ruta_excel", "")
        existe = os.path.exists(ruta) if ruta else False

        card = self._card_base(sv, "📤  Actualizar datos de reporte (mensual)")
        self._fila_excel(card, p, existe, ruta)

        if existe:
            n_rep, ult = self._info_reporte_str(ruta)
            lbl = f"📅  Reporte actual: {n_rep} mes(es) — último: {ult}" if n_rep > 0 \
                  else "📅  Sin datos de reporte aún"
            ctk.CTkLabel(card, text=lbl, font=(FONTS.family, FONTS.size_sm),
                         text_color=COLORS.text_secondary).pack(anchor="w", padx=16, pady=(0, 4))

        self._sep_card(card)

        ctk.CTkLabel(card, text="¿Cuántos meses nuevos quieres agregar?",
                     font=(FONTS.family, FONTS.size_sm, "bold"),
                     text_color=COLORS.text_primary).pack(anchor="w", padx=16, pady=(0, 6))

        fila = ctk.CTkFrame(card, fg_color="transparent")
        fila.pack(fill="x", padx=16, pady=(0, 4))
        self.entry_meses = ctk.CTkEntry(fila, width=80, height=34,
                                        font=(FONTS.family, FONTS.size_md, "bold"),
                                        justify="center")
        self.entry_meses.insert(0, "6")
        self.entry_meses.pack(side="left", padx=(0, 8))
        ctk.CTkLabel(fila, text="meses nuevos",
                     font=(FONTS.family, FONTS.size_sm),
                     text_color=COLORS.text_secondary).pack(side="left")

        self.lbl_preview = ctk.CTkLabel(card, text="",
                                        font=(FONTS.family, FONTS.size_xs),
                                        text_color=COLORS.primary)
        self.lbl_preview.pack(anchor="w", padx=16, pady=(0, 6))
        self.entry_meses.bind("<KeyRelease>",
                              lambda e: self._preview_mensual(p, existe, ruta))
        if existe:
            self._preview_mensual(p, existe, ruta)

        self.lbl_descarga = ctk.CTkLabel(card, text="",
                                         font=(FONTS.family, FONTS.size_xs),
                                         text_color=COLORS.success,
                                         wraplength=560, justify="left")
        self.lbl_descarga.pack(anchor="w", padx=16)

        ctk.CTkButton(card, text="⬇  Descargar Excel con filas nuevas",
                      font=(FONTS.family, FONTS.size_sm, "bold"),
                      fg_color=COLORS.primary, hover_color=COLORS.primary_hover,
                      height=36, corner_radius=6,
                      command=lambda: self._descargar_mensual(p, existe, ruta)
                      ).pack(anchor="w", padx=16, pady=(6, 10))

        self._seccion_subir(card, p, existe, ruta)

    def _preview_mensual(self, p, existe, ruta):
        if not existe:
            self.lbl_preview.configure(text=""); return
        try:
            n = int(self.entry_meses.get())
            if n < 1 or n > 60: raise ValueError()
        except ValueError:
            self.lbl_preview.configure(
                text="  ⚠  Ingresa un número entre 1 y 60", text_color=COLORS.error)
            return
        freq   = "mensual"
        ultimo = _ultimo_periodo_reporte(ruta, freq) or _ultimo_periodo_base(ruta, freq)
        if not ultimo:
            self.lbl_preview.configure(text=""); return
        fechas = _generar_fechas_desde(ultimo, n, freq)
        self.lbl_preview.configure(
            text=f"  → {_fmt_fecha(fechas[0], freq)}  a  {_fmt_fecha(fechas[-1], freq)}  ({n} meses nuevos)",
            text_color=COLORS.primary)

    def _descargar_mensual(self, p, existe, ruta):
        if not existe:
            self.lbl_descarga.configure(
                text="⚠  Busca primero el Excel del proyecto.", text_color=COLORS.error)
            return
        try:
            n = int(self.entry_meses.get())
            if n < 1 or n > 60: raise ValueError()
        except ValueError:
            self.lbl_descarga.configure(
                text="⚠  Ingresa un número de meses válido (1–60).", text_color=COLORS.error)
            return
        freq   = "mensual"
        ultimo = _ultimo_periodo_reporte(ruta, freq) or _ultimo_periodo_base(ruta, freq)
        if not ultimo:
            self.lbl_descarga.configure(
                text="⚠  No se pudo leer la fecha del Excel.", text_color=COLORS.error); return

        nuevas_fechas = _generar_fechas_desde(ultimo, n, freq)
        self._ejecutar_descarga(p, ruta, nuevas_fechas, freq)

    # ══════════════════════════════════════════════════════════════════════════
    # PANEL DIARIO
    # ══════════════════════════════════════════════════════════════════════════
    def _panel_actualizar_diario(self, p: dict):
        sv   = self.right_sv
        ruta = p.get("ruta_excel", "")
        existe = os.path.exists(ruta) if ruta else False
        freq   = "diario"

        card = self._card_base(sv, "📤  Actualizar datos de reporte (datos diarios)")
        self._fila_excel(card, p, existe, ruta)

        if existe:
            n_rep, ult = self._info_reporte_str(ruta)
            lbl = f"📅  Reporte actual: {n_rep} día(s) — último: {ult}" if n_rep > 0 \
                  else "📅  Sin datos de reporte aún"
            ctk.CTkLabel(card, text=lbl, font=(FONTS.family, FONTS.size_sm),
                         text_color=COLORS.text_secondary).pack(anchor="w", padx=16, pady=(0, 4))

        self._sep_card(card)

        hint_frame = ctk.CTkFrame(card, fg_color="#EBF5FB", corner_radius=6)
        hint_frame.pack(fill="x", padx=16, pady=(0, 8))
        ctk.CTkLabel(hint_frame,
                     text="💡  Ingresa el rango de días que quieres agregar al reporte.\n"
                          "    Formato: dd/mm/yyyy  (ej: 01/04/2025)",
                     font=(FONTS.family, FONTS.size_xs),
                     text_color="#1B4F72",
                     justify="left").pack(anchor="w", padx=10, pady=6)

        ctk.CTkLabel(card, text="Rango de días nuevos a agregar:",
                     font=(FONTS.family, FONTS.size_sm, "bold"),
                     text_color=COLORS.text_primary).pack(anchor="w", padx=16, pady=(0, 4))

        fila_rango = ctk.CTkFrame(card, fg_color="transparent")
        fila_rango.pack(fill="x", padx=16, pady=(0, 4))

        ctk.CTkLabel(fila_rango, text="Desde:",
                     font=(FONTS.family, FONTS.size_sm),
                     text_color=COLORS.text_secondary, width=60).pack(side="left")
        self.entry_desde_d = ctk.CTkEntry(fila_rango, width=120, height=34,
                                           font=(FONTS.family, FONTS.size_sm),
                                           justify="center",
                                           placeholder_text="dd/mm/yyyy")
        self.entry_desde_d.pack(side="left", padx=(4, 16))

        ctk.CTkLabel(fila_rango, text="Hasta:",
                     font=(FONTS.family, FONTS.size_sm),
                     text_color=COLORS.text_secondary, width=60).pack(side="left")
        self.entry_hasta_d = ctk.CTkEntry(fila_rango, width=120, height=34,
                                           font=(FONTS.family, FONTS.size_sm),
                                           justify="center",
                                           placeholder_text="dd/mm/yyyy")
        self.entry_hasta_d.pack(side="left", padx=(4, 0))

        if existe:
            ultimo = _ultimo_periodo_reporte(ruta, freq) or _ultimo_periodo_base(ruta, freq)
            if ultimo:
                siguiente = ultimo + timedelta(days=1)
                self.entry_desde_d.insert(0, _fmt_fecha(siguiente, freq))

        self.lbl_preview_d = ctk.CTkLabel(card, text="",
                                           font=(FONTS.family, FONTS.size_xs),
                                           text_color=COLORS.primary)
        self.lbl_preview_d.pack(anchor="w", padx=16, pady=(0, 6))

        def _actualizar_preview_d(*_):
            self._preview_rango_diario()

        self.entry_desde_d.bind("<KeyRelease>", _actualizar_preview_d)
        self.entry_hasta_d.bind("<KeyRelease>", _actualizar_preview_d)

        self.lbl_descarga = ctk.CTkLabel(card, text="",
                                         font=(FONTS.family, FONTS.size_xs),
                                         text_color=COLORS.success,
                                         wraplength=560, justify="left")
        self.lbl_descarga.pack(anchor="w", padx=16)

        ctk.CTkButton(card, text="⬇  Descargar Excel con filas nuevas",
                      font=(FONTS.family, FONTS.size_sm, "bold"),
                      fg_color=COLORS.primary, hover_color=COLORS.primary_hover,
                      height=36, corner_radius=6,
                      command=lambda: self._descargar_diario(p, existe, ruta)
                      ).pack(anchor="w", padx=16, pady=(6, 10))

        self._seccion_subir(card, p, existe, ruta)

    def _preview_rango_diario(self):
        try:
            d = _parsear_date_str(self.entry_desde_d.get())
            h = _parsear_date_str(self.entry_hasta_d.get())
            if h < d:
                raise ValueError("'Hasta' debe ser posterior a 'Desde'")
            n = (h - d).days + 1
            self.lbl_preview_d.configure(
                text=f"  → {_fmt_fecha(d, 'diario')} a {_fmt_fecha(h, 'diario')} — {n} días nuevos",
                text_color=COLORS.primary)
        except ValueError as e:
            self.lbl_preview_d.configure(
                text=f"  ⚠  {e}" if "Hasta" in str(e) else "  (completa ambas fechas en formato dd/mm/yyyy)",
                text_color=COLORS.text_secondary)

    def _descargar_diario(self, p, existe, ruta):
        if not existe:
            self.lbl_descarga.configure(
                text="⚠  Busca primero el Excel del proyecto.", text_color=COLORS.error); return
        try:
            d = _parsear_date_str(self.entry_desde_d.get())
            h = _parsear_date_str(self.entry_hasta_d.get())
            if h < d:
                raise ValueError("'Hasta' debe ser posterior o igual a 'Desde'")
        except ValueError as e:
            self.lbl_descarga.configure(
                text=f"⚠  Fechas inválidas: {e}", text_color=COLORS.error); return
        nuevas_fechas = _generar_fechas_rango(d, h, "diario")
        self._ejecutar_descarga(p, ruta, nuevas_fechas, "diario")

    # ══════════════════════════════════════════════════════════════════════════
    # PANEL HORARIO
    # ══════════════════════════════════════════════════════════════════════════
    def _panel_actualizar_horario(self, p: dict):
        sv   = self.right_sv
        ruta = p.get("ruta_excel", "")
        existe = os.path.exists(ruta) if ruta else False
        freq   = "horario"

        card = self._card_base(sv, "📤  Actualizar datos de reporte (datos horarios)")
        self._fila_excel(card, p, existe, ruta)

        if existe:
            n_rep, ult = self._info_reporte_str(ruta)
            lbl = f"📅  Reporte actual: {n_rep} hora(s) — último: {ult}" if n_rep > 0 \
                  else "📅  Sin datos de reporte aún"
            ctk.CTkLabel(card, text=lbl, font=(FONTS.family, FONTS.size_sm),
                         text_color=COLORS.text_secondary).pack(anchor="w", padx=16, pady=(0, 4))

        self._sep_card(card)

        hint_frame = ctk.CTkFrame(card, fg_color="#EBF5FB", corner_radius=6)
        hint_frame.pack(fill="x", padx=16, pady=(0, 8))
        ctk.CTkLabel(hint_frame,
                     text="💡  Ingresa el rango horario que quieres agregar al reporte.\n"
                          "    Formato fecha: dd/mm/yyyy    Hora: número entre 0 y 23",
                     font=(FONTS.family, FONTS.size_xs),
                     text_color="#1B4F72",
                     justify="left").pack(anchor="w", padx=10, pady=6)

        ctk.CTkLabel(card, text="Desde (fecha y hora de inicio):",
                     font=(FONTS.family, FONTS.size_sm, "bold"),
                     text_color=COLORS.text_primary).pack(anchor="w", padx=16, pady=(0, 4))

        fila_desde = ctk.CTkFrame(card, fg_color="transparent")
        fila_desde.pack(fill="x", padx=16, pady=(0, 6))

        ctk.CTkLabel(fila_desde, text="Fecha:",
                     font=(FONTS.family, FONTS.size_sm),
                     text_color=COLORS.text_secondary, width=55).pack(side="left")
        self.entry_desde_hf = ctk.CTkEntry(fila_desde, width=120, height=34,
                                            font=(FONTS.family, FONTS.size_sm),
                                            justify="center", placeholder_text="dd/mm/yyyy")
        self.entry_desde_hf.pack(side="left", padx=(4, 12))
        ctk.CTkLabel(fila_desde, text="Hora:",
                     font=(FONTS.family, FONTS.size_sm),
                     text_color=COLORS.text_secondary, width=45).pack(side="left")
        self.entry_desde_hh = ctk.CTkEntry(fila_desde, width=60, height=34,
                                            font=(FONTS.family, FONTS.size_sm),
                                            justify="center", placeholder_text="0-23")
        self.entry_desde_hh.pack(side="left", padx=(4, 0))

        ctk.CTkLabel(card, text="Hasta (fecha y hora de fin, inclusive):",
                     font=(FONTS.family, FONTS.size_sm, "bold"),
                     text_color=COLORS.text_primary).pack(anchor="w", padx=16, pady=(0, 4))

        fila_hasta = ctk.CTkFrame(card, fg_color="transparent")
        fila_hasta.pack(fill="x", padx=16, pady=(0, 6))

        ctk.CTkLabel(fila_hasta, text="Fecha:",
                     font=(FONTS.family, FONTS.size_sm),
                     text_color=COLORS.text_secondary, width=55).pack(side="left")
        self.entry_hasta_hf = ctk.CTkEntry(fila_hasta, width=120, height=34,
                                            font=(FONTS.family, FONTS.size_sm),
                                            justify="center", placeholder_text="dd/mm/yyyy")
        self.entry_hasta_hf.pack(side="left", padx=(4, 12))
        ctk.CTkLabel(fila_hasta, text="Hora:",
                     font=(FONTS.family, FONTS.size_sm),
                     text_color=COLORS.text_secondary, width=45).pack(side="left")
        self.entry_hasta_hh = ctk.CTkEntry(fila_hasta, width=60, height=34,
                                            font=(FONTS.family, FONTS.size_sm),
                                            justify="center", placeholder_text="0-23")
        self.entry_hasta_hh.pack(side="left", padx=(4, 0))

        if existe:
            ultimo = _ultimo_periodo_reporte(ruta, freq) or _ultimo_periodo_base(ruta, freq)
            if ultimo:
                siguiente = ultimo + timedelta(hours=1)
                self.entry_desde_hf.insert(0, siguiente.strftime("%d/%m/%Y"))
                self.entry_desde_hh.insert(0, str(siguiente.hour))

        self.lbl_preview_h = ctk.CTkLabel(card, text="",
                                           font=(FONTS.family, FONTS.size_xs),
                                           text_color=COLORS.primary)
        self.lbl_preview_h.pack(anchor="w", padx=16, pady=(0, 6))

        def _actualizar_preview_h(*_):
            self._preview_rango_horario()

        for entry in (self.entry_desde_hf, self.entry_desde_hh,
                      self.entry_hasta_hf, self.entry_hasta_hh):
            entry.bind("<KeyRelease>", _actualizar_preview_h)

        self.lbl_descarga = ctk.CTkLabel(card, text="",
                                         font=(FONTS.family, FONTS.size_xs),
                                         text_color=COLORS.success,
                                         wraplength=560, justify="left")
        self.lbl_descarga.pack(anchor="w", padx=16)

        ctk.CTkButton(card, text="⬇  Descargar Excel con filas nuevas",
                      font=(FONTS.family, FONTS.size_sm, "bold"),
                      fg_color=COLORS.primary, hover_color=COLORS.primary_hover,
                      height=36, corner_radius=6,
                      command=lambda: self._descargar_horario(p, existe, ruta)
                      ).pack(anchor="w", padx=16, pady=(6, 10))

        self._seccion_subir(card, p, existe, ruta)

    def _preview_rango_horario(self):
        try:
            d = _parsear_datetime_str(self.entry_desde_hf.get(), self.entry_desde_hh.get())
            h = _parsear_datetime_str(self.entry_hasta_hf.get(), self.entry_hasta_hh.get())
            if h < d:
                raise ValueError("'Hasta' debe ser posterior a 'Desde'")
            n_horas = int((h - d).total_seconds() / 3600) + 1
            self.lbl_preview_h.configure(
                text=f"  → {_fmt_fecha(d, 'horario')}  a  {_fmt_fecha(h, 'horario')}  — {n_horas} horas nuevas",
                text_color=COLORS.primary)
        except ValueError as e:
            self.lbl_preview_h.configure(
                text=f"  ⚠  {e}" if "Hasta" in str(e) else "  (completa todos los campos)",
                text_color=COLORS.text_secondary)

    def _descargar_horario(self, p, existe, ruta):
        if not existe:
            self.lbl_descarga.configure(
                text="⚠  Busca primero el Excel del proyecto.", text_color=COLORS.error); return
        try:
            d = _parsear_datetime_str(self.entry_desde_hf.get(), self.entry_desde_hh.get())
            h = _parsear_datetime_str(self.entry_hasta_hf.get(), self.entry_hasta_hh.get())
            if h < d:
                raise ValueError("'Hasta' debe ser posterior o igual a 'Desde'")
        except ValueError as e:
            self.lbl_descarga.configure(
                text=f"⚠  Datos inválidos: {e}", text_color=COLORS.error); return
        nuevas_fechas = _generar_fechas_rango(d, h, "horario")
        self._ejecutar_descarga(p, ruta, nuevas_fechas, "horario")

    # ══════════════════════════════════════════════════════════════════════════
    # HELPERS COMPARTIDOS
    # ══════════════════════════════════════════════════════════════════════════

    def _card_base(self, sv, titulo: str) -> ctk.CTkFrame:
        card = ctk.CTkFrame(sv, fg_color=COLORS.bg_card,
                             corner_radius=10, border_width=2,
                             border_color=COLORS.primary)
        card.pack(fill="x", padx=4, pady=(0, 4))
        ctk.CTkLabel(card, text=titulo,
                     font=(FONTS.family, FONTS.size_md, "bold"),
                     text_color=COLORS.primary
                     ).pack(anchor="w", padx=16, pady=(14, 4))
        return card

    def _fila_excel(self, card, p, existe, ruta):
        icon  = "✅" if existe else "⚠️"
        color = COLORS.success if existe else COLORS.error
        fila  = ctk.CTkFrame(card, fg_color="transparent")
        fila.pack(fill="x", padx=16, pady=(0, 4))
        nombre_xls = os.path.basename(ruta) if ruta else "No definido"
        ctk.CTkLabel(fila, text=f"{icon}  Excel del proyecto: {nombre_xls}",
                     font=(FONTS.family, FONTS.size_sm, "bold"),
                     text_color=color).pack(side="left")
        if not existe:
            ctk.CTkButton(fila, text="Buscar",
                          font=(FONTS.family, FONTS.size_xs),
                          fg_color=COLORS.primary, height=26, width=70,
                          command=lambda: self._rebuscar_excel(p)
                          ).pack(side="left", padx=8)

    def _sep_card(self, card):
        ctk.CTkFrame(card, fg_color=COLORS.border, height=1
                     ).pack(fill="x", padx=16, pady=(4, 10))

    def _seccion_subir(self, card, p, existe, ruta):
        """Sección inferior de subir Excel — igual para todas las frecuencias."""
        self._sep_card(card)

        ctk.CTkLabel(card, text="📂  Subir Excel con los nuevos datos llenados",
                     font=(FONTS.family, FONTS.size_sm, "bold"),
                     text_color=COLORS.text_primary, anchor="w"
                     ).pack(anchor="w", padx=16, pady=(0, 4))

        ctk.CTkLabel(card,
                     text=(
                         "Llena los valores nuevos en la hoja 'Reporte' y sube el archivo.\n"
                         "El Excel del proyecto se actualizará y los gráficos se recalcularán."
                     ),
                     font=(FONTS.family, FONTS.size_sm),
                     text_color=COLORS.text_secondary, justify="left"
                     ).pack(anchor="w", padx=16, pady=(0, 8))

        self.lbl_subida_ok = ctk.CTkLabel(card, text="",
                                           font=(FONTS.family, FONTS.size_xs),
                                           text_color=COLORS.success)
        self.lbl_subida_ok.pack(anchor="w", padx=16)

        self.progress_bar = ctk.CTkProgressBar(card, mode="indeterminate",
                                               height=8, corner_radius=4,
                                               progress_color=COLORS.accent)
        self.progress_bar.pack(fill="x", padx=16, pady=(4, 0))
        self.progress_bar.pack_forget()

        self.lbl_subida_err = ctk.CTkLabel(card, text="",
                                            font=(FONTS.family, FONTS.size_xs),
                                            text_color=COLORS.error,
                                            wraplength=560, justify="left")
        self.lbl_subida_err.pack(anchor="w", padx=16)

        btns = ctk.CTkFrame(card, fg_color="transparent")
        btns.pack(fill="x", padx=16, pady=(6, 14))

        ctk.CTkButton(btns, text="📂  Subir y actualizar proyecto",
                      font=(FONTS.family, FONTS.size_sm, "bold"),
                      text_color=COLORS.primary_dark,
                      fg_color=COLORS.accent, hover_color=COLORS.accent_hover,
                      height=38, corner_radius=6,
                      command=lambda: self._subir_y_calcular(p, existe, ruta)
                      ).pack(side="left", padx=(0, 8))

        if existe:
            ctk.CTkButton(btns, text="🔄  Abrir proyecto",
                          font=(FONTS.family, FONTS.size_sm),
                          fg_color="transparent",
                          hover_color=COLORS.primary_light,
                          text_color=COLORS.primary,
                          border_width=1, border_color=COLORS.primary,
                          height=38, corner_radius=6,
                          command=lambda: self._calcular(p, ruta)
                          ).pack(side="left")

    # ── Descarga genérica ─────────────────────────────────────────────────────
    def _ejecutar_descarga(self, p, ruta, nuevas_fechas, frecuencia: str):
        nombre_base = os.path.splitext(os.path.basename(ruta))[0]
        sugerido    = f"{nombre_base}_actualizado.xlsx"
        path_dest   = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=sugerido,
            title="Guardar Excel con filas nuevas",
        )
        if not path_dest:
            return
        try:
            from data.plantilla import expandir_reporte
            n_exist, n_new = expandir_reporte(
                path_origen=ruta,
                path_destino=path_dest,
                nuevas_fechas=nuevas_fechas,
                col_consumo=p.get("col_consumo", "Consumo_kWh"),
                vars_independientes=p.get("vars_independientes", []),
                unidad=p.get("unidad_energia", "kWh"),
                frecuencia=frecuencia,
            )
        except Exception as e:
            self.lbl_descarga.configure(
                text=f"❌  Error al generar el archivo: {e}", text_color=COLORS.error)
            return

        nombre = os.path.basename(path_dest)
        unidad_tiempo = {"mensual": "mes(es)", "diario": "día(s)", "horario": "hora(s)"}
        ut = unidad_tiempo.get(frecuencia, "período(s)")
        if n_new > 0:
            self.lbl_descarga.configure(
                text=(f"✅  {nombre}\n"
                      f"     {n_exist} {ut} existentes conservados  +  {n_new} {ut} nuevos agregados\n"
                      f"     Llena los valores nuevos en la hoja 'Reporte' y súbelo abajo."),
                text_color=COLORS.success)
        else:
            self.lbl_descarga.configure(
                text=f"ℹ️  {nombre}  —  No se agregaron períodos (todos ya existían).",
                text_color=COLORS.text_secondary)

    # ── Info reporte ──────────────────────────────────────────────────────────
    def _info_reporte_str(self, ruta: str):
        """Devuelve (n_filas, ultimo_como_str)."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            if "Reporte" not in wb.sheetnames:
                wb.close(); return 0, "—"
            ws = wb["Reporte"]
            ultima = None; count = 0
            for row in ws.iter_rows(min_row=4, max_col=1, values_only=True):
                if row[0] is not None and str(row[0]).strip():
                    ultima = str(row[0]).strip(); count += 1
            wb.close()
            return count, (ultima or "—")
        except Exception:
            return 0, "—"

    # ─────────────────────────────────────────────────────────────────────────
    # SUBIR Y CALCULAR / CALCULAR
    # ─────────────────────────────────────────────────────────────────────────
    def _rebuscar_excel(self, p: dict):
        path = filedialog.askopenfilename(
            filetypes=[("Excel", "*.xlsx *.xls")],
            title="Buscar Excel del proyecto",
        )
        if not path:
            return
        from data.gestor_proyectos import actualizar_ruta_excel
        actualizar_ruta_excel(p["_archivo"], path)
        p["ruta_excel"] = path
        self._seleccionar(p)

    def _subir_y_calcular(self, p: dict, existe: bool, ruta_actual: str):
        if not existe:
            self.lbl_subida_err.configure(
                text="⚠  El Excel base no se encontró. Búscalo primero.")
            return
        path_nuevo = filedialog.askopenfilename(
            filetypes=[("Excel", "*.xlsx *.xls")],
            title="Seleccionar Excel con datos actualizados",
        )
        if not path_nuevo:
            return
        self.lbl_subida_err.configure(text="")
        if os.path.normpath(path_nuevo) != os.path.normpath(ruta_actual):
            from data.gestor_proyectos import actualizar_ruta_excel
            actualizar_ruta_excel(p["_archivo"], path_nuevo)
            p["ruta_excel"] = path_nuevo
            ruta_actual = path_nuevo
        self._calcular(p, ruta_actual)

    def _calcular(self, p: dict, ruta: str):
        from data.lector_excel import leer_excel
        from data.validador import validar_dataframe
        from core.calculadora import calcular

        col_consumo = p.get("col_consumo", "Consumo_kWh")
        vars_ind    = p.get("vars_independientes", [])
        modelo_id   = p.get("modelo_id", "promedio")
        nivel_conf  = p.get("nivel_confianza", 95)
        frecuencia  = _frecuencia_proyecto(p)

        try:
            df_hist = leer_excel(ruta, hoja="Base")
            err_h = validar_dataframe(df_hist, col_consumo, vars_ind)
            if err_h:
                self.lbl_subida_err.configure(text="❌ Base — " + " | ".join(err_h))
                return

            df_rep = None
            try:
                df_rep_raw = leer_excel(ruta, hoja="Reporte")
                df_rep = df_rep_raw.dropna(subset=[col_consumo]).reset_index(drop=True)
                if len(df_rep) == 0:
                    df_rep = None
                    self.lbl_subida_ok.configure(
                        text="✅  Archivo cargado — sin datos de reporte llenados todavía.",
                        text_color=COLORS.text_secondary)
                else:
                    err_r = validar_dataframe(df_rep, col_consumo, vars_ind)
                    if err_r:
                        self.lbl_subida_err.configure(
                            text="❌ Hoja Reporte — " + " | ".join(err_r))
                        return
                    ut = {"mensual": "mes(es)", "diario": "día(s)", "horario": "hora(s)"}.get(frecuencia, "período(s)")
                    self.lbl_subida_ok.configure(
                        text=f"✅  {len(df_rep)} {ut} de reporte cargados",
                        text_color=COLORS.success)
                    self.progress_bar.pack(fill="x", padx=16, pady=(4, 4))
                    self.progress_bar.start()
                    self.update_idletasks()
            except Exception:
                pass

            resultado = calcular(df_hist, df_rep, modelo_id,
                                  col_consumo, vars_ind, nivel_conf,
                                  frecuencia=frecuencia)
            self.progress_bar.stop()
            self.progress_bar.pack_forget()
            resultado["unidad"] = p.get("unidad_energia", "kWh")
            self._resultado = resultado

            s = self.app.sesion
            s.nombre_proyecto     = p.get("nombre_proyecto", "")
            s.unidad_energia      = p.get("unidad_energia", "kWh")
            s.modelo_id           = modelo_id
            s.col_consumo         = col_consumo
            s.vars_independientes = vars_ind
            s.nivel_confianza     = nivel_conf
            s.df_base        = df_hist
            s.df_reporte          = df_rep
            s.resultado           = resultado

            self.app.show_page("ResultadosPage", desde="MonitoreoPage")

        except Exception as exc:
            self.progress_bar.stop()
            self.progress_bar.pack_forget()
            self.lbl_subida_err.configure(text=f"❌ Error: {exc}")