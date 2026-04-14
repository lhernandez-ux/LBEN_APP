"""
core/exportador.py
==================
Exporta el informe LBEn en Excel — solo con datos que realmente existen
en `resultado` y `sesion`, espejando lo que muestra la UI.

Hojas:
  1. Portada        — identificación básica del proyecto
  2. Línea_Base     — tabla_lben_completa (lo que muestra Tab 1)
  3. Desempeño      — tabla_desempeno     (lo que muestra Tab 2)
  4. Seguimiento    — CUSUM + datos de reporte (lo que muestra Tab 3)
  5. Ajuste_NR      — solo si hay_anr=True (lo que muestra Tab 4)
"""

from __future__ import annotations
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from data.sesion import Sesion


# ── Paleta (formato aRGB 8 dígitos, sin #, requerido por openpyxl) ───────────
_C_DARK       = "FF1B3A6B"
_C_MID        = "FF2563A8"
_C_LIGHT      = "FFEBF2FB"
_C_GREEN      = "FF1A6B3C"
_C_GREEN_DARK = "FF1A6B3C"   # para banda resumen seguimiento
_C_GREEN_L    = "FFE9F7EF"
_C_RED_L      = "FFFDECEC"
_C_AMBER      = "FF7D5A0A"
_C_AMBER_L    = "FFFEF9E7"
_C_WHITE      = "FFFFFFFF"
_C_ALT        = "FFF4F6FA"
_C_RED_FONT   = "FFBF2020"   # rojo para fuente (degradación)

_THIN  = Side(style="thin",   color="FFC5D0E0")
_MED   = Side(style="medium", color="FF2563A8")
_THICK = Side(style="medium", color="FF1B3A6B")
_B     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_B_MED = Border(left=_MED,  right=_MED,  top=_MED,  bottom=_MED)

_AC = Alignment(horizontal="center", vertical="center", wrap_text=False)
_AL = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_AR = Alignment(horizontal="right",  vertical="center", wrap_text=False)


def _fill(c: str):
    """PatternFill asegurado: convierte 6 dígitos a 8 (aRGB) si es necesario."""
    c = c.lstrip("#")
    if len(c) == 6:
        c = "FF" + c.upper()
    return PatternFill("solid", fgColor=c.upper())


def _font(bold=False, color="FF222222", size=10, italic=False):
    # Normalizar color de fuente también
    color = color.lstrip("#")
    if len(color) == 6:
        color = "FF" + color.upper()
    return Font(name="Arial", bold=bold, color=color.upper(), size=size, italic=italic)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _banda(ws, texto, n_cols, row, bg=_C_DARK, fg=_C_WHITE, size=11, height=24):
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws.cell(row=row, column=1, value=texto)
    c.fill      = _fill(bg)
    c.font      = _font(bold=True, color=fg, size=size)
    c.alignment = _AC
    c.border    = _B_MED
    ws.row_dimensions[row].height = height


def _encabezados(ws, columnas, row=2, height=20):
    fills = [_C_DARK] + [_C_MID] * (len(columnas) - 1)
    for j, (col, fill) in enumerate(zip(columnas, fills), start=1):
        c = ws.cell(row=row, column=j, value=col)
        c.fill      = _fill(fill)
        c.font      = _font(bold=True, color=_C_WHITE, size=10)
        c.alignment = _AC
        c.border    = _B
    ws.row_dimensions[row].height = height


def _fila_datos(ws, row, valores, alt=False):
    bg = _C_ALT if alt else _C_WHITE
    for j, val in enumerate(valores, start=1):
        c = ws.cell(row=row, column=j, value=val)
        c.fill      = _fill(bg)
        c.border    = _B
        c.alignment = _AL if j == 1 else _AR
        c.font      = _font(size=10)
    ws.row_dimensions[row].height = 17


def _auto_ancho(ws, n_cols, min_w=12, max_w=38):
    for j in range(1, n_cols + 1):
        col_letter = get_column_letter(j)
        max_len = 0
        for row in ws.iter_rows(min_col=j, max_col=j):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 3, max_w))


def _borde_ext(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            L = _THICK if c == c1 else _THIN
            R = _THICK if c == c2 else _THIN
            T = _THICK if r == r1 else _THIN
            B = _THICK if r == r2 else _THIN
            cell.border = Border(left=L, right=R, top=T, bottom=B)


def _nota_pie(ws, n_cols, row, texto):
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws.cell(row=row, column=1, value=texto)
    c.font      = _font(italic=True, color=_C_DARK, size=9)
    c.fill      = _fill(_C_LIGHT)
    c.alignment = _AL


# ── Entrada pública ───────────────────────────────────────────────────────────

def exportar_informe(path: str, resultado: dict, sesion: "Sesion"):
    if path.endswith(".pdf"):
        _exportar_pdf(path, resultado, sesion)
    else:
        _exportar_excel(path, resultado, sesion)


# ── Excel ─────────────────────────────────────────────────────────────────────

def _exportar_excel(path: str, resultado: dict, sesion: "Sesion"):
    import tempfile, os, shutil

    wb = openpyxl.Workbook()
    _hoja_portada(wb, resultado, sesion)
    _hoja_linea_base(wb, resultado, sesion)
    _hoja_desempeno(wb, resultado, sesion)
    _hoja_seguimiento(wb, resultado, sesion)
    if resultado.get("hay_anr"):
        _hoja_ajuste_nr(wb, resultado, sesion)
    wb.active = wb["Portada"]

    # Guardar en temporal primero para evitar archivos corruptos
    dir_destino = os.path.dirname(path) or "."
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=dir_destino) as tmp:
        tmp_path = tmp.name
    try:
        wb.save(tmp_path)
        shutil.move(tmp_path, path)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 1 — Portada
# ══════════════════════════════════════════════════════════════════════════════

def _hoja_portada(wb, resultado, sesion):
    ws = wb.active
    ws.title = "Portada"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 52

    cabeceras = [
        ("UNIDAD DE PLANEACIÓN MINERO ENERGÉTICA — UPME",            _C_DARK,  _C_WHITE, 14, 32),
        ("Informe de Línea Base de Consumo Energético (LBEn)",        _C_MID,   _C_WHITE, 12, 24),
        ("Resolución UPME 16 de 2024 — Art. 30, Ley 1715/2014",      _C_LIGHT, _C_DARK,  10, 18),
    ]
    for row, (txt, bg, fg, size, h) in enumerate(cabeceras, start=1):
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=txt)
        c.fill      = _fill(bg)
        c.font      = _font(bold=(row < 3), italic=(row == 3), color=fg, size=size)
        c.alignment = _AC
        ws.row_dimensions[row].height = h

    kpis      = resultado.get("kpis", {})
    modelo_id = getattr(sesion, "modelo_id", "promedio")
    fechas    = resultado.get("fechas", [])

    ws.merge_cells("A5:B5")
    ws["A5"] = "IDENTIFICACIÓN DEL PROYECTO"
    ws["A5"].fill      = _fill(_C_MID)
    ws["A5"].font      = _font(bold=True, color=_C_WHITE, size=10)
    ws["A5"].alignment = _AC
    ws.row_dimensions[5].height = 20

    ficha = [
        ("Proyecto / Entidad",     getattr(sesion, "nombre_proyecto", None) or "—"),
        ("Zona climática",         getattr(sesion, "zona_climatica",   None) or "—"),
        ("Energético / Unidad",    getattr(sesion, "unidad_energia",   "kWh")),
        ("Modelo LBEn",            _modelo_nombre(modelo_id)),
        ("Período base",           _rango_fechas(fechas)),
        ("N.° períodos base",      str(kpis.get("n_periodos", "—"))),
        ("Fecha de generación",    _hoy()),
        ("Normativa",              "Res. UPME 16/2024 · ISO 50006:2014 · ISO 50001:2018"),
    ]

    for i, (k, v) in enumerate(ficha, start=6):
        ws.row_dimensions[i].height = 20
        ka = ws.cell(row=i, column=1, value=k)
        ka.font = _font(bold=True, color=_C_DARK, size=10)
        ka.fill = _fill(_C_LIGHT); ka.border = _B; ka.alignment = _AL
        va = ws.cell(row=i, column=2, value=v)
        va.font = _font(size=10); va.border = _B; va.alignment = _AL

    fila_kpi = 6 + len(ficha) + 1
    ws.merge_cells(f"A{fila_kpi}:B{fila_kpi}")
    ws.cell(row=fila_kpi, column=1, value="MÉTRICAS DEL MODELO").fill = _fill(_C_DARK)
    ws.cell(row=fila_kpi, column=1).font      = _font(bold=True, color=_C_WHITE, size=10)
    ws.cell(row=fila_kpi, column=1).alignment = _AC
    ws.row_dimensions[fila_kpi].height = 20

    unidad = getattr(sesion, "unidad_energia", "kWh")
    kpi_filas = _construir_kpis_portada(kpis, modelo_id, unidad)

    for i, (k, v) in enumerate(kpi_filas, start=fila_kpi + 1):
        ws.row_dimensions[i].height = 19
        ka = ws.cell(row=i, column=1, value=k)
        ka.font = _font(bold=True, color=_C_DARK, size=10)
        ka.fill = _fill(_C_LIGHT); ka.border = _B; ka.alignment = _AL
        va = ws.cell(row=i, column=2, value=v)
        va.font = _font(size=10); va.border = _B; va.alignment = _AL

    _borde_ext(ws, 5, 1, fila_kpi + len(kpi_filas), 2)


def _construir_kpis_portada(kpis: dict, modelo_id: str, unidad: str) -> list:
    rows = []

    def _add(label, key, fmt=None, suffix=""):
        val = kpis.get(key)
        if val is not None:
            txt = (fmt.format(val) if fmt else str(val)) + (f" {suffix}" if suffix else "")
            rows.append((label, txt))

    _add("N.° períodos utilizados",       "n_periodos")
    _add("SEM — Error estándar",          "sem",          "{:,.2f}", unidad)
    _add("CV — Coeficiente de variación", "cv",           "{:.2f}%")

    if modelo_id in ("promedio", "cociente"):
        _add("LBEn promedio anual", "lben_promedio_anual", "{:,.2f}", unidad)

    if modelo_id == "regresion":
        _add("R²",            "r2",           "{:.4f}")
        _add("R² ajustado",   "r2_ajustado",  "{:.4f}")
        _add("CV(RMSE)",      "cv_rmse",      "{:.2f}%")
        _add("F-estadístico", "f_estadistico","{:.2f}")
        _add("p-valor F",     "p_valor_f",    "{:.4f}")

    return rows


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 2 — Línea Base
# ══════════════════════════════════════════════════════════════════════════════

def _hoja_linea_base(wb, resultado, sesion):
    ws = wb.create_sheet("Línea_Base")
    ws.sheet_view.showGridLines = False

    tabla    = resultado.get("tabla_lben_completa", [])
    columnas = resultado.get("cols_lben_completa",  [])

    if not tabla or not columnas:
        unidad   = getattr(sesion, "unidad_energia", "kWh")
        columnas = ["Período",
                    f"Consumo real ({unidad})",
                    f"LBEn ({unidad})",
                    f"IC superior ({unidad})",
                    f"IC inferior ({unidad})",
                    "Outlier"]
        tabla = [
            [str(f), _fmt_num(r), _fmt_num(b), "—", "—", ""]
            for f, r, b in zip(
                resultado.get("fechas", []),
                resultado.get("consumo_real", []),
                resultado.get("linea_base", []),
            )
        ]

    n = len(columnas)
    _banda(ws,
           f"Línea Base de Consumo Energético (LBEn)  |  {getattr(sesion, 'nombre_proyecto', '') or ''}",
           n, 1, bg=_C_DARK)
    _banda(ws, "Art. 7.5 — Resolución UPME 16/2024",
           n, 2, bg=_C_MID, size=9, height=18)
    _encabezados(ws, columnas, row=3)

    for i, fila in enumerate(tabla):
        r      = i + 4
        alt    = i % 2 == 0
        es_out = len(fila) > 5 and str(fila[5]) not in ("", "—", "0", "False")
        bg_row = _C_RED_L if es_out else (_C_ALT if alt else _C_WHITE)

        ws.row_dimensions[r].height = 17
        for j, val in enumerate(fila, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.fill      = _fill(bg_row)
            c.border    = _B
            c.alignment = _AL if j == 1 else _AR
            c.font      = _font(size=10, bold=(j == 1))

    ws.freeze_panes = "B4"
    _auto_ancho(ws, n)
    _borde_ext(ws, 1, 1, 3 + len(tabla), n)

    r_nota = 4 + len(tabla) + 1
    _nota_pie(ws, n, r_nota,
              "Outlier = valor fuera del IC depurado. "
              "IC = LBEn ± 10% (Ec. 3, Art. 7.5.1, Res. UPME 16/2024).")

    if getattr(sesion, "modelo_id", "") == "regresion":
        params = resultado.get("modelo_params", {})
        if params:
            _seccion_diagnosticos_regresion(ws, params, r_nota + 2,
                                            getattr(sesion, "unidad_energia", "kWh"))


def _seccion_diagnosticos_regresion(ws, params: dict, row_ini: int, unidad: str):
    coefs = params.get("coeficientes", {})
    if not coefs:
        return

    _banda(ws, "Diagnósticos del Modelo de Regresión Lineal", 6, row_ini, bg=_C_DARK)
    r = row_ini + 1

    vars_ind = [k for k in coefs if k != "Intercepto"]
    partes   = [f"{coefs.get('Intercepto', 0):,.2f}"]
    for var in vars_ind:
        v = coefs[var]
        signo = "+" if v >= 0 else "−"
        partes.append(f"{signo} {abs(v):,.4f}·{var}")
    ec_txt = "LBEn = " + "  ".join(partes)

    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value=ec_txt)
    c.fill = _fill(_C_LIGHT); c.border = _B
    c.font = _font(bold=True, color=_C_DARK, size=10)
    c.alignment = _AL
    ws.row_dimensions[r].height = 20
    r += 1

    def _met(label, key, fmt="{}", nota=""):
        val = params.get(key)
        if val is not None:
            return (label, fmt.format(val), nota)
        return None

    metricas = [
        _met("R²",            "r2",           "{:.4f}",  "≥ 0.75 recomendado"),
        _met("R² ajustado",   "r2_ajustado",  "{:.4f}",  "Penaliza variables innecesarias"),
        _met("RMSE",          "rmse",         "{:,.2f}", unidad),
        _met("CV(RMSE)",      "cv_rmse",      "{:.2f}%", "≤ 20% recomendado"),
        _met("F-estadístico", "f_estadistico","{:.2f}",  ""),
        _met("p-valor F",     "p_valor_f",    "{:.4f}",  "< 0.05 para modelo significativo"),
        _met("N períodos",    "n",            "{}",      ""),
    ]
    metricas = [m for m in metricas if m]

    _encabezados(ws, ["Métrica", "Valor", "Referencia"], row=r)
    r += 1
    for i, (lbl, val, nota) in enumerate(metricas):
        for j, txt in enumerate([lbl, val, nota], start=1):
            c = ws.cell(row=r, column=j, value=txt)
            c.fill      = _fill(_C_ALT if i % 2 == 0 else _C_WHITE)
            c.border    = _B
            c.alignment = _AL
            c.font      = _font(bold=(j == 1), color=_C_DARK if j == 1 else "FF222222", size=10)
        ws.row_dimensions[r].height = 18
        r += 1

    p_vals    = params.get("p_valores", {})
    pearson_r = params.get("pearson_r", {})
    vif       = params.get("vif", {})
    vif_cols  = bool(vif)
    cols_var  = ["Variable", "Coeficiente", "p-valor", "r Pearson"]
    if vif_cols:
        cols_var.append("VIF")

    r += 1
    _banda(ws, "Coeficientes y significancia por variable", len(cols_var), r, bg=_C_MID, size=9, height=18)
    r += 1
    _encabezados(ws, cols_var, row=r)
    r += 1

    for i, var in enumerate(vars_ind):
        pv  = p_vals.get(var)
        rv  = pearson_r.get(var)
        vv  = vif.get(var) if vif_cols else None
        sig = ("***" if pv and pv < 0.001
               else "**" if pv and pv < 0.01
               else "*"  if pv and pv < 0.05
               else "✗"  if pv is not None else "—")
        fila = [
            var,
            f"{coefs[var]:,.4f}",
            f"{pv:.4f} {sig}" if pv is not None else "—",
            f"{rv:+.2f}"      if rv is not None else "—",
        ]
        if vif_cols:
            fila.append(f"{vv:.2f}" if vv is not None else "—")

        for j, val in enumerate(fila, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.fill      = _fill(_C_ALT if i % 2 == 0 else _C_WHITE)
            c.border    = _B
            c.alignment = _AL if j == 1 else _AR
            c.font      = _font(size=10)
        ws.row_dimensions[r].height = 17
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 3 — Desempeño
# ══════════════════════════════════════════════════════════════════════════════

def _hoja_desempeno(wb, resultado, sesion):
    ws = wb.create_sheet("Desempeño")
    ws.sheet_view.showGridLines = False

    columnas = resultado.get("columnas_desempeno", [])
    filas    = resultado.get("tabla_desempeno",    [])

    if not columnas or not filas:
        ws.cell(row=1, column=1, value="Sin datos de desempeño disponibles")
        return

    n = len(columnas)
    _banda(ws,
           f"Tabla de Desempeño Energético  |  {getattr(sesion, 'nombre_proyecto', '') or ''}",
           n, 1, bg=_C_DARK)
    _banda(ws, "Ahorro = E_base − E_reporte ± ajustes  (Ec. 12/13, Resolución UPME 16/2024)",
           n, 2, bg=_C_MID, size=9, height=18)
    _encabezados(ws, columnas, row=3)

    try:
        di = columnas.index("Desviación (%)") + 1
    except ValueError:
        di = -1

    for i, fila in enumerate(filas):
        r   = i + 4
        alt = i % 2 == 0
        bg  = _C_ALT if alt else _C_WHITE

        if di > 0 and di <= len(fila):
            try:
                v = float(str(fila[di-1]).replace("%","").replace("+","").replace(",",".").strip())
                if   v < -2: bg = _C_GREEN_L
                elif v > 2:  bg = _C_RED_L
            except ValueError:
                pass

        ws.row_dimensions[r].height = 17
        for j, val in enumerate(fila, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.fill      = _fill(bg)
            c.border    = _B
            c.alignment = _AL if j == 1 else _AR
            c.font      = _font(size=10, bold=(j == 1))
            if j == di:
                try:
                    v = float(str(val).replace("%","").replace("+","").replace(",",".").strip())
                    c.font = _font(bold=True, size=10,
                                   color=_C_GREEN if v <= 0 else _C_RED_FONT)
                except ValueError:
                    pass

    ws.freeze_panes = "B4"
    _auto_ancho(ws, n)
    _borde_ext(ws, 1, 1, 3 + len(filas), n)

    r_ley = 4 + len(filas) + 1
    _nota_pie(ws, n, r_ley,
              "Verde: desviación ≤ −2% (ahorro)   |   "
              "Rojo: desviación ≥ +2% (exceso)   |   "
              "Blanco/gris: dentro de ±2%")


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 4 — Seguimiento (CUSUM + reporte)
# ══════════════════════════════════════════════════════════════════════════════

def _hoja_seguimiento(wb, resultado, sesion):
    ws = wb.create_sheet("Seguimiento")
    ws.sheet_view.showGridLines = False
    unidad = getattr(sesion, "unidad_energia", "kWh")

    fechas   = resultado.get("fechas",         [])
    desv_abs = resultado.get("desviacion_abs",  [])
    desv_pct = resultado.get("desviacion_pct",  [])
    cusum    = resultado.get("cusum",           [])

    if not fechas:
        ws.cell(row=1, column=1, value="Sin datos de seguimiento disponibles")
        return

    tiene_reporte = resultado.get("tiene_reporte", False)

    columnas = ["Período",
                f"Desviación absoluta ({unidad})",
                "Desviación (%)",
                f"CUSUM acumulado ({unidad})",
                "Estado"]
    n = len(columnas)

    _banda(ws,
           f"Seguimiento Energético — CUSUM  |  {getattr(sesion, 'nombre_proyecto', '') or ''}",
           n, 1, bg=_C_DARK)
    _banda(ws, "Acumulado de desviaciones respecto a la LBEn  |  Resolución UPME 16/2024",
           n, 2, bg=_C_MID, size=9, height=18)

    if tiene_reporte and desv_pct:
        excesos = sum(1 for d in desv_pct if d > 0)
        ahorros = sum(1 for d in desv_pct if d <= 0)
        prom    = sum(desv_pct) / len(desv_pct)
        # CORRECTO: usar constante _C_GREEN_DARK (aRGB 8 dígitos), nunca string con #
        _banda(ws,
               f"Resumen: {ahorros} mes(es) con ahorro  |  {excesos} sobre LBEn  |  "
               f"Desviación promedio: {prom:+.1f}%",
               n, 3, bg=_C_GREEN_DARK, fg=_C_WHITE, size=9, height=18)
        fila_enc = 4
    else:
        fila_enc = 3

    _encabezados(ws, columnas, row=fila_enc)

    n_filas = max(len(fechas), len(desv_abs), len(cusum))
    for i in range(n_filas):
        r   = i + fila_enc + 1
        alt = i % 2 == 0

        f  = fechas[i]   if i < len(fechas)   else "—"
        da = desv_abs[i] if i < len(desv_abs) else None
        dp = desv_pct[i] if i < len(desv_pct) else None
        cu = cusum[i]    if i < len(cusum)     else None

        estado = ""
        bg     = _C_ALT if alt else _C_WHITE
        if da is not None:
            if da < 0:
                estado = "Mejor que LBEn"
                bg     = _C_GREEN_L
            elif da > 0:
                estado = "Peor que LBEn"
                bg     = _C_RED_L
            else:
                estado = "En LBEn"

        fila = [
            str(f),
            _fmt_num(da) if da is not None else "—",
            f"{dp:+.2f}%" if dp is not None else "—",
            _fmt_num(cu)  if cu is not None else "—",
            estado,
        ]

        ws.row_dimensions[r].height = 17
        for j, val in enumerate(fila, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.fill      = _fill(bg)
            c.border    = _B
            c.alignment = _AL if j == 1 else _AR
            c.font      = _font(size=10, bold=(j == 1))
            if j == 4 and cu is not None:
                c.font = _font(bold=True, size=10,
                               color=_C_GREEN if cu <= 0 else _C_RED_FONT)

    r_fin = fila_enc + 1 + n_filas
    ws.freeze_panes = "B" + str(fila_enc + 1)
    _auto_ancho(ws, n)
    _borde_ext(ws, 1, 1, r_fin - 1, n)

    _nota_pie(ws, n, r_fin,
              "CUSUM acumulado negativo indica ahorro sostenido. "
              "CUSUM creciente indica deterioro del desempeño energético.")


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 5 — Ajuste No Rutinario (ANR)
# ══════════════════════════════════════════════════════════════════════════════

def _hoja_ajuste_nr(wb, resultado, sesion):
    ws = wb.create_sheet("Ajuste_NR")
    ws.sheet_view.showGridLines = False
    unidad  = getattr(sesion, "unidad_energia", "kWh")
    resumen = resultado.get("resumen_anr", {})
    detalle = resumen.get("detalle", [])

    columnas = ["Período", "Motivo del ajuste",
                f"Valor original ({unidad})",
                f"Valor ajustado ({unidad})",
                "Δ (%)", "Ajustado"]
    n = len(columnas)

    _banda(ws, "Ajustes No Rutinarios (ANR) — Art. 7.6.2, Resolución UPME 16/2024",
           n, 1, bg=_C_DARK)
    _banda(ws, "Factores estáticos que afectan el consumo fuera del modelo",
           n, 2, bg=_C_AMBER, fg=_C_WHITE, size=9, height=18)

    n_marc = resumen.get("n_meses_marcados", 0)
    n_adj  = resumen.get("n_ajustados",      0)
    n_no   = resumen.get("n_no_ajustados",   0)
    años   = resumen.get("años_afectados",   [])
    anios_txt = ", ".join(str(a) for a in años) if años else "—"

    _banda(ws,
           f"Meses marcados: {n_marc}  |  Ajustados: {n_adj}  |  "
           f"Sin ajuste posible: {n_no}  |  Años afectados: {anios_txt}",
           n, 3, bg=_C_AMBER_L, fg=_C_AMBER, size=9, height=18)

    _encabezados(ws, columnas, row=4)

    for i, rec in enumerate(detalle):
        r        = i + 5
        ajustado = rec.get("ajustado", False)
        delta    = rec.get("delta_pct", None)
        bg       = _C_ALT if i % 2 == 0 else _C_WHITE
        if not ajustado:
            bg = _C_AMBER_L

        v_orig = rec.get("valor_original", "—")
        v_adj  = rec.get("valor_ajustado", "—") if ajustado else "Sin ajuste"

        fila = [
            str(rec.get("fecha",  "—")),
            str(rec.get("motivo", "—")),
            _fmt_num(v_orig) if isinstance(v_orig, (int, float)) else str(v_orig),
            _fmt_num(v_adj)  if isinstance(v_adj,  (int, float)) else str(v_adj),
            f"{delta:+.1f}%" if delta is not None else "—",
            "Sí" if ajustado else "No",
        ]
        ws.row_dimensions[r].height = 17
        for j, val in enumerate(fila, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.fill      = _fill(bg)
            c.border    = _B
            c.alignment = _AL if j <= 2 else _AR
            c.font      = _font(size=10, bold=(j == 1))

    ws.freeze_panes = "B5"
    _auto_ancho(ws, n)
    _borde_ext(ws, 1, 1, 4 + len(detalle), n)

    _nota_pie(ws, n, 5 + len(detalle),
              "ANR = Ajuste No Rutinario. "
              "Ámbar = mes marcado pero sin ajuste posible (sin referencia disponible).")


# ── PDF ───────────────────────────────────────────────────────────────────────

def _exportar_pdf(path: str, resultado: dict, sesion: "Sesion"):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer, HRFlowable,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
    except ImportError:
        raise ImportError("Instala reportlab: pip install reportlab")

    doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=1.8*cm, bottomMargin=1.8*cm)
    styles = getSampleStyleSheet()
    # reportlab sí acepta hex de 6 dígitos con #
    azul  = colors.HexColor("#1B3A6B")
    azul2 = colors.HexColor("#2563A8")
    gris  = colors.HexColor("#F4F6FA")

    h1   = ParagraphStyle("h1",   parent=styles["Title"],
                          textColor=azul,  fontSize=16, spaceAfter=4)
    h2   = ParagraphStyle("h2",   parent=styles["Heading2"],
                          textColor=azul2, fontSize=11, spaceAfter=4, spaceBefore=10)
    nota = ParagraphStyle("nota", parent=styles["Normal"],
                          textColor=colors.HexColor("#555555"),
                          fontSize=8, italic=True)

    kpis   = resultado.get("kpis", {})
    unidad = getattr(sesion, "unidad_energia", "kWh")
    modelo = _modelo_nombre(getattr(sesion, "modelo_id", "promedio"))
    fechas = resultado.get("fechas", [])

    story = [
        Paragraph("UNIDAD DE PLANEACIÓN MINERO ENERGÉTICA — UPME", h1),
        Paragraph("Informe de Línea Base de Consumo Energético (LBEn)", h2),
        Paragraph("Resolución UPME 16 de 2024", nota),
        Spacer(1, 0.3*cm),
        HRFlowable(width="100%", thickness=1.5, color=azul),
        Spacer(1, 0.3*cm),
    ]

    ficha = [
        ["Proyecto / Entidad",  getattr(sesion, "nombre_proyecto", "—") or "—"],
        ["Unidad energética",   unidad],
        ["Modelo LBEn",         modelo],
        ["Período base",        _rango_fechas(fechas)],
        ["N.° períodos",        str(kpis.get("n_periodos", "—"))],
        ["Fecha",               _hoy()],
    ]
    story.append(Paragraph("Identificación del Proyecto", h2))
    story.append(_pdf_tabla(ficha, col_widths=[6*cm, 12*cm],
                            header=False, azul=azul, gris=gris))
    story.append(Spacer(1, 0.4*cm))

    kpi_rows = [["Indicador", "Valor"]]
    def _add_kpi(lbl, key, fmt="{}", suffix=""):
        v = kpis.get(key)
        if v is not None:
            kpi_rows.append([lbl, fmt.format(v) + (f" {suffix}" if suffix else "")])

    _add_kpi("N.° períodos", "n_periodos")
    _add_kpi("SEM",          "sem",     "{:,.2f}", unidad)
    _add_kpi("CV",           "cv",      "{:.2f}%")
    modelo_id = getattr(sesion, "modelo_id", "promedio")
    if modelo_id == "regresion":
        _add_kpi("R²",       "r2",      "{:.4f}")
        _add_kpi("CV(RMSE)", "cv_rmse", "{:.2f}%")

    if len(kpi_rows) > 1:
        story.append(Paragraph("Métricas del Modelo", h2))
        story.append(_pdf_tabla(kpi_rows, col_widths=[8*cm, 6*cm],
                                header=True, azul=azul, gris=gris))
        story.append(Spacer(1, 0.4*cm))

    cols_d = resultado.get("columnas_desempeno", [])
    fils_d = resultado.get("tabla_desempeno",    [])
    if cols_d and fils_d:
        story.append(Paragraph("Tabla de Desempeño Energético", h2))
        col_w = [max(3*cm, 18*cm / len(cols_d))] * len(cols_d)
        story.append(_pdf_tabla([cols_d] + fils_d, col_widths=col_w,
                                header=True, azul=azul, gris=gris))

    doc.build(story)


def _pdf_tabla(data, col_widths, header=True, azul=None, gris=None):
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    azul = azul or colors.HexColor("#1B3A6B")
    gris = gris or colors.HexColor("#F4F6FA")
    t = Table(data, colWidths=col_widths, repeatRows=1 if header else 0)
    style = [
        ("FONTNAME",       (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",       (0, 0), (-1, -1), 9),
        ("GRID",           (0, 0), (-1, -1), 0.4, colors.HexColor("#C5D0E0")),
        ("ROWBACKGROUNDS", (0, 1 if header else 0), (-1, -1), [gris, colors.white]),
        ("ALIGN",          (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN",          (0, 0), (0, -1),  "LEFT"),
        ("TOPPADDING",     (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 4),
    ]
    if header:
        style += [
            ("BACKGROUND", (0, 0), (-1, 0), azul),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ]
    t.setStyle(TableStyle(style))
    return t


# ── Utilidades ────────────────────────────────────────────────────────────────

def _modelo_nombre(modelo_id: str) -> str:
    return {
        "promedio":  "Valor Absoluto de Energía (Art. 7.4.1)",
        "cociente":  "Cociente de Valores Medidos (Art. 7.4.2)",
        "regresion": "Modelo Estadístico — Regresión Lineal (Art. 7.4.3)",
    }.get(modelo_id, modelo_id.title())


def _rango_fechas(fechas: list) -> str:
    if not fechas:
        return "—"
    return f"{fechas[0]}  →  {fechas[-1]}"


def _fmt_num(val) -> str:
    if val is None:
        return "—"
    try:
        return f"{float(val):,.2f}"
    except (TypeError, ValueError):
        return str(val)


def _hoy() -> str:
    from datetime import date
    d = date.today()
    meses = ["enero","febrero","marzo","abril","mayo","junio",
             "julio","agosto","septiembre","octubre","noviembre","diciembre"]
    return f"{d.day} de {meses[d.month-1]} de {d.year}"